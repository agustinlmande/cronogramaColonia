import streamlit as st
import pandas as pd
import random
import io
import openpyxl

# --- CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(page_title="Master Plan Colonia", page_icon="☀️", layout="wide")

# --- 1. CONFIGURACIÓN ---
HORARIOS_ACTIVOS = ["09:00", "09:30", "10:00", "10:30", "11:00", "11:30"]
HORARIOS_VISTA = ["08:30", "09:00", "09:30", "10:00", "10:30", "11:00", "11:30", "12:00"]
DIAS_SEMANA = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES"]

CAPACIDAD_PILETAS = {
    "CHICA": 1,
    "MEDIANA": 3,
    "GRANDE": 3
}

GRUPOS_CON_PROFE_COMPARTIDO = ["CELESTE", "AMARILLO"] 

# Actividades tranquilas permitidas pegadas a la pileta
ACTIVIDADES_PERMITIDAS_BUFFER = ["MERIENDA", "PLAZA"]

# MERIENDA es la unica que permite multiples grupos a la vez
ACTIVIDADES_SIN_PROFE_EXCLUSIVO = ["MERIENDA"]

# --- 2. FUNCIONES DE LÓGICA ---

def extraer_colores_excel(archivo_upload):
    mapa_colores = {}
    try:
        wb = openpyxl.load_workbook(archivo_upload, data_only=True)
        ws = wb.active
        col_idx = None
        for cell in ws[1]:
            if cell.value and "GRUPO" in str(cell.value).upper():
                col_idx = cell.column
                break
        if col_idx:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                val = cell.value
                if val:
                    grupo_nombre = str(val).strip().upper()
                    color_obj = cell.fill.start_color
                    if color_obj.type == 'rgb' and color_obj.rgb:
                        hex_code = color_obj.rgb
                        if len(hex_code) > 6: hex_code = hex_code[2:]
                        mapa_colores[grupo_nombre] = f"#{hex_code}"
    except: pass
    return mapa_colores

def parsear_dias(texto):
    if pd.isna(texto): return []
    texto = str(texto).upper().replace(".", "").strip()
    dias_map = {"LUN": 0, "MAR": 1, "MIE": 2, "MIÉ": 2, "JUE": 3, "VIE": 4}
    if " A " in texto:
        partes = texto.split(" A ")
        inicio, fin = -1, -1
        for k, v in dias_map.items():
            if k in partes[0]: inicio = v
            if k in partes[1]: fin = v
        if inicio != -1 and fin != -1: return list(range(inicio, fin + 1))
    resultado = set()
    texto_limpio = texto.replace(",", " ").replace(" Y ", " ").replace('"', '')
    for palabra in texto_limpio.split():
        for k, v in dias_map.items():
            if k in palabra: resultado.add(v)
    return list(resultado)

def procesar_datos(df):
    df['GRUPO'] = df['GRUPO'].ffill()
    df = df.dropna(subset=['DEPORTE'])
    df['GRUPO'] = df['GRUPO'].astype(str).str.strip().str.upper()
    df['ACTIVIDAD'] = df['DEPORTE'].astype(str).str.strip().str.upper()
    df['TIPO_PILETA'] = df['PILETA'].astype(str).str.upper().str.strip().replace('NAN', '')
    return df

def es_grupo_profe_compartido(nombre_grupo):
    for key in GRUPOS_CON_PROFE_COMPARTIDO:
        if key in nombre_grupo: return True
    return False

def es_compatible_con_natacion(actividad):
    for permitida in ACTIVIDADES_PERMITIDAS_BUFFER:
        if permitida in actividad:
            return True
    return False

def verificar_y_asignar(schedule, ocupacion_grupo, conteo_piletas, ocupacion_recurso, registro_actividades_diarias,
                        dia_idx, dia_nom, hora, grupo, actividad, lugar_nombre, es_pileta=False, capacidad_pileta=0):
    
    # 1. UNICIDAD
    actividades_hoy = registro_actividades_diarias.get((dia_nom, grupo), set())
    if actividad in actividades_hoy:
        return False

    # 2. DISPONIBILIDAD GRUPO
    if (dia_nom, hora, grupo) in ocupacion_grupo:
        return False
    
    # 3. DISPONIBILIDAD RECURSO
    if es_pileta:
        key_pileta = (dia_nom, hora, lugar_nombre)
        ocupantes_actuales = conteo_piletas.get(key_pileta, 0)
        if ocupantes_actuales >= capacidad_pileta:
            return False
        conteo_piletas[key_pileta] = ocupantes_actuales + 1
    else:
        if actividad not in ACTIVIDADES_SIN_PROFE_EXCLUSIVO:
            if (dia_nom, hora, actividad) in ocupacion_recurso:
                return False
            ocupacion_recurso.add((dia_nom, hora, actividad))

    # 4. ASIGNAR
    schedule.append({
        "DIA_ID": dia_idx, "DIA": dia_nom, "HORA": hora,
        "GRUPO": grupo, "ACTIVIDAD": actividad, "LUGAR": lugar_nombre
    })
    ocupacion_grupo.add((dia_nom, hora, grupo))
    
    if (dia_nom, grupo) not in registro_actividades_diarias:
        registro_actividades_diarias[(dia_nom, grupo)] = set()
    registro_actividades_diarias[(dia_nom, grupo)].add(actividad)
    
    return True

def generar_cronograma_logica(df_input):
    df = procesar_datos(df_input)
    schedule = []
    
    ocupacion_grupo = set()
    conteo_piletas = {}
    ocupacion_recurso = set()
    registro_actividades_diarias = {}
    mapa_natacion_indices = {} 

    # --- FASE 1: NATACIÓN ---
    df_nat = df[df['ACTIVIDAD'].str.contains("NATAC")]
    grupos_natacion = list(df_nat.iterrows())
    random.shuffle(grupos_natacion)

    for _, row in grupos_natacion:
        grupo = row['GRUPO']
        p_raw = row['TIPO_PILETA']
        capacidad = 3; lugar = "PILETA GRANDE"
        if "CHICA" in p_raw: capacidad = 1; lugar = "PILETA CHICA"
        elif "MEDIANA" in p_raw: capacidad = 3; lugar = "PILETA MEDIANA"
        
        necesita_profe = es_grupo_profe_compartido(grupo)
        asignado_fijo = False
        posibles_horarios = list(HORARIOS_ACTIVOS)
        random.shuffle(posibles_horarios)
        
        for hora in posibles_horarios:
            es_viable_toda_semana = True
            for dia_nom in DIAS_SEMANA:
                if conteo_piletas.get((dia_nom, hora, lugar), 0) >= capacidad:
                    es_viable_toda_semana = False; break
                if (dia_nom, hora, grupo) in ocupacion_grupo:
                    es_viable_toda_semana = False; break
                if necesita_profe and (dia_nom, hora, "PROFE_NATACION_MENORES") in ocupacion_recurso:
                    es_viable_toda_semana = False; break
            
            if es_viable_toda_semana:
                idx_hora_elegida = HORARIOS_ACTIVOS.index(hora)
                for d_idx, dia_nom in enumerate(DIAS_SEMANA):
                    verificar_y_asignar(schedule, ocupacion_grupo, conteo_piletas, ocupacion_recurso, registro_actividades_diarias,
                                        d_idx, dia_nom, hora, grupo, "NATACION", lugar, True, 99)
                    if necesita_profe: ocupacion_recurso.add((dia_nom, hora, "PROFE_NATACION_MENORES"))
                    mapa_natacion_indices[(dia_nom, grupo)] = idx_hora_elegida
                asignado_fijo = True
                break
        
        if not asignado_fijo:
            st.toast(f"⚠️ {grupo} sin natación.", icon="🚨")

    # --- FASE 2: OTROS DEPORTES ---
    df_otros = df[~df['ACTIVIDAD'].str.contains("NATAC")].copy()
    df_otros['PRIORIDAD'] = df_otros['ACTIVIDAD'].apply(lambda x: 2 if any(k in x for k in ["TENIS","RIO","CANOTAJE","HOCKEY","PLAZA"]) else 1)
    df_otros = df_otros.sort_values(by='PRIORIDAD', ascending=False)
    
    for _, row in df_otros.iterrows():
        grupo = row['GRUPO']
        actividad = row['ACTIVIDAD']
        dias_posibles = parsear_dias(row['DÍAS'])
        try: frec = int(str(row['ESTÍMULO']).split()[0])
        except: frec = 1
        
        lugar = "CANCHA"
        if "TENIS" in actividad: lugar = "TENIS"
        elif "CANOTAJE" in actividad: lugar = "RIO"
        elif "HOCKEY" in actividad: lugar = "CANCHA HOCKEY"
        elif "PLAZA" in actividad: lugar = "PLAZA"
        
        es_merienda_o_plaza = es_compatible_con_natacion(actividad)

        asignados = 0; intentos = 0
        while asignados < frec and intentos < 200:
            intentos += 1
            if not dias_posibles: break
            
            carga = {}
            for d in dias_posibles:
                n = DIAS_SEMANA[d]
                c = sum(1 for s in schedule if s['GRUPO'] == grupo and s['DIA'] == n)
                carga[d] = c
            dias_ord = sorted(dias_posibles, key=lambda d: (carga[d], random.random()))
            
            asignado_ahora = False
            for dia_idx in dias_ord:
                dia_nom = DIAS_SEMANA[dia_idx]
                if actividad in registro_actividades_diarias.get((dia_nom, grupo), set()): continue 

                hs = list(HORARIOS_ACTIVOS)
                random.shuffle(hs)
                idx_natacion = mapa_natacion_indices.get((dia_nom, grupo))
                
                for hora in hs:
                    idx_actual = HORARIOS_ACTIVOS.index(hora)
                    
                    if not es_merienda_o_plaza and idx_natacion is not None:
                        if abs(idx_actual - idx_natacion) == 1: continue 

                    if verificar_y_asignar(schedule, ocupacion_grupo, conteo_piletas, ocupacion_recurso, registro_actividades_diarias,
                                           dia_idx, dia_nom, hora, grupo, actividad, lugar, False, 0):
                        asignados += 1; asignado_ahora = True; break
                if asignado_ahora: break

    # --- FASE 3: RELLENO ---
    todos_grupos = df['GRUPO'].unique()
    for g in todos_grupos:
        for d_idx, d_nom in enumerate(DIAS_SEMANA):
            for h in HORARIOS_VISTA:
                if (d_nom, h, g) not in ocupacion_grupo:
                    schedule.append({ "DIA_ID": d_idx, "DIA": d_nom, "HORA": h, "GRUPO": g, "ACTIVIDAD": "LIBRE", "LUGAR": "-" })

    return pd.DataFrame(schedule)

# --- 3. MODAL DE PAUTAS ---
@st.dialog("📋 Pautas de Generación")
def mostrar_pautas():
    st.markdown("""
    Estas son las reglas matemáticas que el sistema respeta para crear el cronograma:
    
    1.  **NATACIÓN FIJA:** El horario de pileta asignado se mantiene igual de Lunes a Viernes.
    2.  **CAPACIDAD PILETAS:**
        * **Chica:** 1 grupo máx.
        * **Mediana/Grande:** 3 grupos máx.
    3.  **PROFE COMPARTIDO:** Los grupos **Celestes y Amarillos** nunca se superponen en natación (comparten profe). Los Naranjas son independientes.
    4.  **UNICIDAD:** Un grupo no repite la misma actividad el mismo día (Ej: No puede tener 2 veces Fútbol el Martes).
    5.  **ZONA DE DESCANSO:** No se asignan deportes intensos inmediatamente antes o después de la pileta.
        * *Excepción:* **Merienda y Plaza** sí pueden ir pegados a la pileta.
    6.  **PROFESORES DE CAMPO:** Cada deporte (Fútbol, Básquet, etc.) tiene 1 solo profe. Si un grupo lo usa, el horario se bloquea para el resto.
        * *Excepción:* **Merienda** es la única actividad que pueden realizar varios grupos a la vez.
    7.  **PLAZA:** Se considera un recurso exclusivo (solo 1 grupo a la vez).
    8.  **HORARIOS:** * 08:30 y 12:00: Siempre Libre.
        * 09:00 a 11:30: Actividades.
    """)

# --- INTERFAZ ---
st.title("Master Plan Colonia ☀️")

# CSS para el botón amarillo
st.markdown("""
<style>
    div[data-testid="column"] button[kind="secondary"] {
        border-color: #FFC107 !important;
        color: #FFC107 !important;
    }
    div[data-testid="column"] button[kind="secondary"]:hover {
        border-color: #FFD54F !important;
        color: #FFD54F !important;
        background-color: #FFF8E1 !important;
    }
</style>
""", unsafe_allow_html=True)

archivo = st.file_uploader("Cargar Excel Nuevo", type=['xlsx', 'csv'])

if archivo:
    try:
        mapa_colores = {}
        if not archivo.name.endswith('.csv'):
            mapa_colores = extraer_colores_excel(archivo)
            archivo.seek(0)

        if archivo.name.endswith('.csv'): df_in = pd.read_csv(archivo)
        else: df_in = pd.read_excel(archivo)
        
        # --- BOTONES EN COLUMNAS ---
        col_gen, col_pautas, col_void = st.columns([2, 1, 3])
        
        with col_gen:
            btn_generar = st.button("GENERAR CRONOGRAMA", type="primary", use_container_width=True)
        with col_pautas:
            if st.button("📋 Ver Pautas", type="secondary", use_container_width=True):
                mostrar_pautas()
        
        if btn_generar:
            with st.spinner("Procesando todas las reglas..."):
                st.session_state['df_res'] = generar_cronograma_logica(df_in)
                st.session_state['mapa_colores'] = mapa_colores
                st.success("¡Cronograma Completado!")

        if 'df_res' in st.session_state:
            df_res = st.session_state['df_res']
            colores = st.session_state['mapa_colores']
            st.divider()
            
            # 1. VISTA GRUPOS
            st.subheader("📁 Vista por Grupos")
            grupos = sorted(df_res['GRUPO'].unique())
            g_sel = st.selectbox("Seleccionar Grupo:", grupos)
            df_g = df_res[df_res['GRUPO'] == g_sel]
            pivot_g = df_g.pivot(index='HORA', columns='DIA', values='ACTIVIDAD')
            pivot_g = pivot_g.reindex(HORARIOS_VISTA).reindex(columns=DIAS_SEMANA)
            st.table(pivot_g)
            
            buffer_grupos = io.BytesIO()
            with pd.ExcelWriter(buffer_grupos, engine='xlsxwriter') as writer:
                workbook = writer.book
                fmt_libre = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
                fmt_actividad = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': True})
                
                df_res.sort_values(by=['GRUPO', 'DIA_ID', 'HORA'])[['GRUPO', 'DIA', 'HORA', 'ACTIVIDAD', 'LUGAR']].to_excel(writer, index=False, sheet_name='RESUMEN')
                
                for grupo in grupos:
                    df_fil = df_res[df_res['GRUPO'] == grupo]
                    mat = df_fil.pivot(index='HORA', columns='DIA', values='ACTIVIDAD')
                    mat = mat.reindex(index=HORARIOS_VISTA, columns=DIAS_SEMANA)
                    sn = str(grupo)[:30].replace(":", "").replace("/", "")
                    mat.to_excel(writer, sheet_name=sn)
                    ws = writer.sheets[sn]
                    ws.set_column(0, 0, 15); ws.set_column(1, 5, 20)
                    for r in range(mat.shape[0]):
                        for c in range(mat.shape[1]):
                            val = mat.iloc[r, c]
                            if val == "LIBRE" or pd.isna(val) or val == "-":
                                ws.write(r+1, c+1, val, fmt_libre)
                            else:
                                ws.write(r+1, c+1, val, fmt_actividad)
            st.download_button("📥 Descargar Plan Grupos (.xlsx)", buffer_grupos.getvalue(), "Cronograma_Grupos.xlsx", "application/vnd.ms-excel")

            st.divider()

            # 2. VISTA PROFESORES
            st.subheader("🎓 Vista por Profesores / Deportes")
            st.info("Solo deportes (Sin Merienda, Plaza ni Natación)")
            
            NO_MOSTRAR = ["LIBRE", "-", "MERIENDA", "PLAZA", "NATACION"]
            actividades = sorted([a for a in df_res['ACTIVIDAD'].unique() if a not in NO_MOSTRAR])
            
            if actividades:
                act_sel = st.selectbox("Seleccionar Deporte:", actividades)
                df_p = df_res[df_res['ACTIVIDAD'] == act_sel]
                pivot_p = df_p.pivot_table(index='HORA', columns='DIA', values='GRUPO', aggfunc=lambda x: ' / '.join(x))
                pivot_p = pivot_p.reindex(HORARIOS_VISTA).reindex(columns=DIAS_SEMANA)
                st.table(pivot_p.fillna("-"))
                
                buffer_profes = io.BytesIO()
                with pd.ExcelWriter(buffer_profes, engine='xlsxwriter') as writer:
                    workbook = writer.book
                    fmt_std = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
                    
                    for act in actividades:
                        df_act = df_res[df_res['ACTIVIDAD'] == act]
                        mat = df_act.pivot_table(index='HORA', columns='DIA', values='GRUPO', aggfunc=lambda x: ' / '.join(x))
                        mat = mat.reindex(index=HORARIOS_VISTA, columns=DIAS_SEMANA).fillna("-")
                        sn = str(act)[:30].replace(":", "").replace("/", "")
                        mat.to_excel(writer, sheet_name=sn)
                        ws = writer.sheets[sn]
                        ws.set_column(0, 0, 15); ws.set_column(1, 5, 25)
                        for r in range(mat.shape[0]):
                            for c in range(mat.shape[1]):
                                ws.write(r+1, c+1, mat.iloc[r, c], fmt_std)
                st.download_button("📥 Descargar Plan Profesores (.xlsx)", buffer_profes.getvalue(), "Horarios_Profesores.xlsx", "application/vnd.ms-excel")
            else:
                st.warning("No hay deportes de campo asignados.")
            
            # --- ZONA DE FINALIZACIÓN ---
            st.divider()
            c1, c2, c3 = st.columns([1, 2, 1])
            with c2:
                if st.button("🔄 Reiniciar Todo", type="secondary", use_container_width=True):
                    st.session_state.clear()
                    st.rerun()

    except Exception as e:
        st.error(f"Error: {e}")

# --- MARCA DE AGUA (SIEMPRE VISIBLE AL FINAL) ---
st.markdown("""
    <div style="text-align: right; color: #808080; font-size: 14px; margin-top: 50px; margin-bottom: 20px;">
        Desarrollado por <b>Agustín</b> - Técnico Superior en Desarrollo de Software
    </div>
""", unsafe_allow_html=True)